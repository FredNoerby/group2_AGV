#!/usr/bin/env python
import openpyxl
from product import Product


def load_production_plan():
    """ Goes through the production plan excel sheet and returns a list of products
        """

    # Uses the openpyxl library to open the production plan excel file
    workbook = openpyxl.load_workbook('../Daily_Production_Plan.xlsx')
    type(workbook)

    # Stores the first sheet of the excel file
    sheet_one = workbook.get_sheet_by_name('Plan')
    # Boolean to check if there are more products in the excel file
    more_products = True
    # Incrementer starting from 3 because of layout of excel file
    incrementer = 3
    # Initializes an empty list for storing the products
    production_list = []

    # Adds all the cells that have a value to the production list
    while more_products:
        # If the cell has any value
        if sheet_one.cell(row=incrementer, column=1).value:
            # Creates a object of the product class with the type from the excel file
            production_list.append(Product((incrementer-2), str(sheet_one.cell(row=incrementer, column=1).value)))
        # If there is no value in the cell the boolean is set to false
        else:
            more_products = False
        # Incrementer is incremented with one
        incrementer += 1
    # Returns the list of products
    return production_list


# If this is run as main the cells will be printed
if __name__ == "__main__":
    
    list_of_products = load_production_plan()

    for product in list_of_products:
        print(product.id, product.type, product.parts, product.status)
