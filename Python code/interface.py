#!/usr/bin/env python
# import rospy
# from std_msgs.msg import String
import openpyxl
from product import Product
from robot import Robot
from assembly_line import AssemblyLine
from quality_control import QualityControl
from pickup_spot import PickupSpot
from location import Location
import gc
try:
    import tkinter
except ImportError:  # python 2
    import Tkinter as tkinter


def load_production_plan():
    """ Goes through the production plan excel sheet and returns a list of products
        """

    # Uses the openpyxl library to open the production plan excel file
    workbook = openpyxl.load_workbook('../Daily_Production_Plan.xlsx')
    type(workbook)

    # Stores the first sheet of the excel file
    sheet_one = workbook.get_sheet_by_name('Plan')
    # Boolean to check if there are more products in the excel file
    more_products = True
    # Incrementer starting from 3 because of layout of excel file
    incrementer = 3
    # Initializes an empty list for storing the products
    production_list = []

    # Adds all the cells that have a value to the production list
    while more_products:
        # If the cell has any value
        if sheet_one.cell(row=incrementer, column=1).value:
            # Creates a object of the product class with the type from the excel file
            production_list.append(Product((incrementer-2), str(sheet_one.cell(row=incrementer, column=1).value)))
        # If there is no value in the cell the boolean is set to false
        else:
            more_products = False
        # Incrementer is incremented with one
        incrementer += 1
    # Returns the list of products
    return production_list


def get_next_task(list_of_products):
    """ Goes through a list of products and returns the first incomplete products ID from the list.
        Returns 'none' if all products are done.

        Arguments:
            list_of_products (Product[list]): List of Products to check for next incomplete
        """
    count = 0
    # Goes through the products in the list
    for product in list_of_products:
        # If the product does not have 'Pass' in its status the product is returned
        if "Waiting" in product.status or "Fail" in product.status:
            print("Got product.id " + str(product.id))
            return product
        # Counts products ready for QC
        elif "Ready" in product.status:
            count += 1
    # Returns 'some' if all products are assembled but not all have passed yet
    if count > 0:
        print(str(count) + " waiting to be graded from QC")
        return "some"
    # If all the products in the list have pass in their status 'none' is returned
    return "none"


def get_color(component):
    """ Returns the corresponding color of a specific component

        Arguments:
            component (str): Component to get color for
        """

    # Dictionary of colors
    color_dict = {"C1": ["#07C5F1"],
                  "C2": ["#05cc26"],
                  "C3": ["#f4873e"],
                  "C4": ["#9b65a9"],
                  "C5": ["#f2ca21"],
                  "C6": ["#df4949"]}

    # Checks if the component is in the color dictionary
    if component in color_dict:
        # Returns color if it's in the dictionary
        return color_dict[component]
    else:
        # Else returns a grey color
        return "#e0e0e0"


def start_button_callback(button, window, robot, assembly, list_of_products, component_pickups, interface_status, interface_robot, interface_assembly):
    """ Starts the daily production when the button is pressed
        """

    # Disables the button once it is pressed
    button['state'] = 'disabled'

    # Boolean used to check if all the products for production has been made:
    more_to_make = True
    while more_to_make:
        # Updates interface window
        window.update()
        # Gets first incomplete product ID in list
        first_incomplete = get_next_task(list_of_products)
        # If all are done and passed exit the while loop
        if first_incomplete == "none":
            # Sets boolean more_to_make to False
            more_to_make = False
        # If all products are done but not passed from QC pass
        elif first_incomplete == "some":
            pass
        else:
            # Updates interface window
            window.update()
            # Checks assembly line storage for components
            assembly_missing = assembly.check_storage_for(first_incomplete.components)
            # Enters if the assembly line has all components to make product
            if assembly_missing == "none":
                # Calls assembly line method to assemble product
                assembly.assemble(first_incomplete)
                # Updates the interfaces different sections
                update_interface_products(list_of_products, interface_status)
                update_interface_storage(robot, assembly, interface_robot, interface_assembly)
                # Updates interface window
                window.update()
            else:
                # Updates interface window
                window.update()
                # Check if the robot's storage is full
                if robot.isFull:
                    # Calls method to drive robot to assembly station
                    robot.go_to(assembly.location)
                    # Unloads components at assembly station
                    robot.unload_components(assembly)
                    # Updates the interface
                    update_interface_storage(robot, assembly, interface_robot, interface_assembly)
                    # Updates interface window
                    window.update()
                else:
                    # Check if the robot has some of the missing components
                    robot_missing = robot.check_storage_for(assembly_missing)
                    # If robot has all missing components go to assembly station
                    if robot_missing == "none":
                        # Calls method to drive robot to assembly station
                        robot.go_to(assembly.location)
                        # Unloads components at assembly station
                        robot.unload_components(assembly)
                        # Updates the interface
                        update_interface_storage(robot, assembly, interface_robot, interface_assembly)
                        # Updates interface window
                        window.update()
                    else:
                        spot_to_go = get_pickup_location(robot_missing[0], component_pickups)
                        # Calls method to drive robot to pickup spot
                        robot.go_to(spot_to_go.location)
                        # Removes a component from the pickup spot's storage
                        spot_to_go.remove_component()
                        # Adds the component type to the robot
                        robot.add_to_storage(spot_to_go.type)
                        # Updates the interface
                        update_interface_storage(robot, assembly, interface_robot, interface_assembly)
                        # Updates interface window
                        window.update()

    # Changes interface to display message that production is done
    tkinter.Label(window, text='Daily Production Done. You Can Go Home Now.', width=92)\
        .grid(row=1, column=0, columnspan=12, rowspan=34, sticky='nsew')
    print("Daily production done")
    # Updates interface window
    window.update()


def get_pickup_location(component, component_pickups):
    """ Takes in an component and a list of pickup spots and returns where to pickup the component

        Args:
            component (str): The component to find pickup spot for
            component_pickups (PickupSpot[list]): List from which the location can be found
        """
    # Goes through the pickup spots
    for spot in component_pickups:
        # If the spot stores the component type needed return the spot
        if spot.type == component:
            return spot


def pass_button_callback(quality_c, pass_entry, list_of_products, interface_status):
    """ Passes a product from QC when button is pressed

        Args:
            quality_c (QualityControl): The quality control unit changing the status
            pass_entry (tkinter.Entry): Entry field where the product ID is stored
            list_of_products (Product[list]): The list that contains the product to fail
            interface_status (tkinter.Label[list]): Interface labels to update
        """
    # Stores the input from the entry field as a string
    text_input = pass_entry.get()
    # Clears the entry field
    pass_entry.delete(0, 'end')

    # Goes through all the products to check if one of them matches the ID from the entry field
    for product in list_of_products:
        if str(product.id) == text_input:
            # Call a QualityControl method to fail a product
            quality_c.pass_product(product)
            # Call a method to update the interface
            update_interface_products(list_of_products, interface_status)


def fail_button_callback(quality_c, fail_entry, list_of_products, interface_status):
    """ Fails a product from QC when button is pressed

        Args:
            quality_c (QualityControl): The quality control unit changing the status
            fail_entry (tkinter.Entry): Entry field where the product ID is stored
            list_of_products (Product[list]): The list that contains the product to fail
            interface_status (tkinter.Label[list]): Interface labels to update
        """
    # Stores the input from the entry field as a string
    text_input = fail_entry.get()
    # Clears the entry field
    fail_entry.delete(0, 'end')

    # Goes through all the products to check if one of them matches the ID from the entry field
    for product in list_of_products:
        if str(product.id) == text_input:
            # Call a QualityControl method to fail a product
            quality_c.fail_product(product)
            # Call a method to update the interface
            update_interface_products(list_of_products, interface_status)


def update_interface_products(list_to_show, interface_status):
    """ Updates the interface

        Args:
            list_to_show (list): List of products to show
            interface_status (tkinter.Label[list]): Interface labels to update
        """

    column_counter = 1
    row_counter = 2
    # Goes through the list and adds each product to the interface
    for i in range(len(list_to_show)):
        if i == 35 or i == 70:
            column_counter += 4
            row_counter = 2

        # Checks for the products status and changes the color accordingly
        if "Pass" in list_to_show[i].status:
            interface_status[i].configure(text=list_to_show[i].status, bg='#9ad88f')
        elif "Fail" in list_to_show[i].status:
            interface_status[i].configure(text=list_to_show[i].status, bg='#d68d8e')
        elif list_to_show[i].status == "Ready for QC":
            interface_status[i].configure(text=list_to_show[i].status, bg='#d6ca8b')

        # Increment the row counter by one
        row_counter += 1


def update_interface_storage(robot, assembly, interface_robot, interface_assembly):
    """ Updates the interface

        Args:
            robot (Robot): Robot with storage to show
            assembly (AssemblyLine): Assembly line with storage to show
            interface_robot (tkinter.Label[list]): Interface labels representing robot storage
            interface_assembly (tkinter.Label[list]): Interface labels representing assembly line storage
        """
    # Starts a loop running for two counts
    for i in range(2):
        # Tries to add the robot storage at place i to the interface
        try:
            interface_robot[i].configure(text=robot.storage[i], bg=get_color(robot.storage[i]))
        # If there is nothing at the storage spot the program will return an index error
        except IndexError:
            # It handles the error by setting the robot storage slot as empty in the interface then
            interface_robot[i].configure(text="EMPTY", bg="#e0e0e0")

    # Counters for rows and columns
    row_counter = 7
    column_counter = 13
    # Starts a loop running for two counts
    for i in range(15):
        # Tries to add the assembly line storage at place i to the interface
        try:
            interface_assembly[i].configure(text=assembly.storage[i], bg=get_color(assembly.storage[i]))
        # If there is nothing at the storage spot the program will return an index error
        except IndexError:
            # It handles the error by setting the assembly line storage slot as empty in the interface then
            interface_assembly[i].configure(text="EMPTY", bg="#e0e0e0")
        if column_counter == 14:
            row_counter += 1
            column_counter = 12
        # Increments the column counter
        column_counter += 1


def load_interface(list_of_products, robot, assembly, quality_c, component_pickups):
    """ Creates the main window interface

        Args:
            list_of_products (Product[list]): Products to be loaded into the interface
            robot (Robot): For displaying robot information in the interface
            assembly (AssemblyLine): For displaying the assembly line information in the interface
            quality_c (QualityControl): For failing and passing product from specific quality control unit
            component_pickups (PickupSpot[list]): List of pickup spots
        """
    # Creating a main window for the user interface
    window = tkinter.Tk()
    # Defining title of window
    window.title("Quality Control")
    # Defining the geometry of the window and where on the screen to place it
    window.geometry("{0}x{1}+0+0".format(window.winfo_screenwidth(), window.winfo_screenheight()))

    # Creates a header for the production schedule
    heading = tkinter\
        .Label(window, text='Daily production schedule', relief='ridge', width=30, bg='#2286c3', fg='#ffffff')
    heading.grid(row=0, column=1, columnspan=11, sticky='nsew')

    # Creates sub-headers for products and statuses
    tkinter.Label(window, text='ID', relief='ridge', width=5, bg='#64b5f6')\
        .grid(row=1, column=1, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='ID', relief='ridge', width=5, bg='#64b5f6')\
        .grid(row=1, column=5, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='ID', relief='ridge', width=5, bg='#64b5f6')\
        .grid(row=1, column=9, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Product', relief='ridge', width=10, bg='#64b5f6')\
        .grid(row=1, column=2, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Product', relief='ridge', width=10, bg='#64b5f6')\
        .grid(row=1, column=6, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Product', relief='ridge', width=10, bg='#64b5f6')\
        .grid(row=1, column=10, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Status', relief='ridge', width=15, bg='#64b5f6')\
        .grid(row=1, column=3, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Status', relief='ridge', width=15, bg='#64b5f6')\
        .grid(row=1, column=7, columnspan=1, sticky='nsew')
    tkinter.Label(window, text='Status', relief='ridge', width=15, bg='#64b5f6')\
        .grid(row=1, column=11, columnspan=1, sticky='nsew')

    # rospy.loginfo("Loading in production plan..")
    column_counter = 1
    row_counter = 2

    interface_status = []

    # Goes through the list and adds each product to the interface
    for i in range(len(list_of_products)):
        if i == 35 or i == 70:
            column_counter += 4
            row_counter = 2

        tkinter.Label(window, text=list_of_products[i].id, relief='ridge', width=5, bg='#e0e0e0')\
            .grid(row=row_counter, column=column_counter, columnspan=1, sticky='nsew')
        tkinter.Label(window, text=list_of_products[i].type, relief='ridge', width=10, bg='#e0e0e0')\
            .grid(row=row_counter, column=column_counter + 1, columnspan=1, sticky='nsew')

        interface_status.append(tkinter.Label(window, text=list_of_products[i].status, relief='ridge', width=15, bg='#e0e0e0'))
        interface_status[-1].grid(row=row_counter, column=column_counter + 2, columnspan=1, sticky='nsew')
        row_counter += 1

    # Creates header for robot storage slots
    heading_robot = tkinter.Label(window, text='Robot inventory slots', relief='ridge', width=30, bg='#64b5f6')
    heading_robot.grid(row=3, column=13, columnspan=2, sticky='nsew')
    # Creates list for robot storage labels
    robot_store_labels = []
    # Starts a loop running for two counts
    for i in range(2):
        # Tries to add the robot storage at place i to the interface
        try:
            robot_store_labels.append(tkinter.Label(window, text=robot.storage[i], relief='ridge', width=15, bg=get_color(robot.storage[i])))
            robot_store_labels[-1].grid(row=4, column=(i+13), columnspan=1, sticky='nsew')
        # If there is nothing at the storage spot the program will return an index error
        except IndexError:
            # It handles the error by setting the robot storage slot as empty in the interface then
            robot_store_labels.append(tkinter.Label(window, text="EMPTY", relief='ridge', width=15, bg="#e0e0e0"))
            robot_store_labels[-1].grid(row=4, column=(i+13), columnspan=1, sticky='nsew')

    # Filler to make space between production schedule and slots
    tkinter.Label(window, text=' ', width=1).grid(row=1, column=0, columnspan=1, sticky='nsew')
    tkinter.Label(window, text=' ', width=1).grid(row=1, column=4, columnspan=1, sticky='nsew')
    tkinter.Label(window, text=' ', width=1).grid(row=1, column=8, columnspan=1, sticky='nsew')
    tkinter.Label(window, text=' ', width=1).grid(row=1, column=12, columnspan=1, sticky='nsew')

    # Creates header for assembly line slots
    heading_assembly = tkinter.Label(window, text='Parts in assembly', relief='ridge', width=30, bg='#64b5f6')
    heading_assembly.grid(row=6, column=13, columnspan=2, sticky='nsew')

    # Creates list for assembly storage slots
    assembly_store_labels = []

    row_counter = 7
    column_counter = 13
    # Starts a loop running for 15 counts
    for i in range(15):
        # Tries to add the robot storage at place i to the interface
        try:
            assembly_store_labels.append(tkinter.Label(window, text=assembly.storage[i], relief='ridge', width=15, bg=get_color(assembly.storage[i])))
            assembly_store_labels[-1].grid(row=row_counter, column=column_counter, columnspan=1, sticky='nsew')
        # If there is nothing at the storage spot the program will return an index error
        except IndexError:
            # It handles the error by setting the assembly line storage slot as empty in the interface then
            assembly_store_labels.append(tkinter.Label(window, text="EMPTY", relief='ridge', width=15, bg="#e0e0e0"))
            assembly_store_labels[-1].grid(row=row_counter, column=column_counter, columnspan=1, sticky='nsew')
        if column_counter == 14:
            row_counter += 1
            column_counter = 12
        # Increments the column counter
        column_counter += 1

    # Creating a button for sending production schedule to robot
    start_button = tkinter.Button(window, text='Start Production', command=lambda: start_button_callback(start_button, window, robot, assembly, list_of_products, component_pickups, interface_status, robot_store_labels, assembly_store_labels), relief='ridge', width=15, bg='#504adf', fg='#ffffff', activebackground='#716de1', activeforeground='#ffffff')
    start_button.grid(row=0, column=13, columnspan=2, sticky='nsew')

    # Creating a button for passing a product
    pass_button = tkinter.Button(window, text='Pass Product..', command=lambda: pass_button_callback(quality_c, pass_entry, list_of_products, interface_status), relief='ridge', width=7, bg='#df4a4a', fg='#ffffff', activebackground='#e16d6d', activeforeground='#ffffff')
    pass_button.grid(row=1, column=13, columnspan=1, sticky='nsew')
    # Creating a entry field
    pass_entry = tkinter.Entry(window, text='ID', relief='ridge', width=7)
    pass_entry.grid(row=1, column=14, columnspan=1, sticky='nsew')
    # Creating a button for failing a product
    fail_button = tkinter.Button(window, text='Fail Product..', command=lambda: fail_button_callback(quality_c, fail_entry, list_of_products, interface_status), relief='ridge', width=7, bg='#df4a4a', fg='#ffffff', activebackground='#e16d6d', activeforeground='#ffffff')
    fail_button.grid(row=2, column=13, columnspan=1, sticky='nsew')
    # Creating a entry field
    fail_entry = tkinter.Entry(window, text='ID', relief='ridge', width=7)
    fail_entry.grid(row=2, column=14, columnspan=1, sticky='nsew')

    # Returns the interface window
    return window


def run_interface(window):
    """ Runs the interface window

        Args:
            window (tkinter.Tk): The window to run main loop for
        """
    # For running the tkinter GUI - replaces the rospy.spin function
    window.mainloop()


if __name__ == "__main__":
    gc.disable()

    # Creates a list of type Product with the types from the excel file
    product_list = load_production_plan()

    # Shorter production list for testing purposes
    """product_list = [Product(1, "P1"), Product(2, "P2"), Product(3, "P3"), Product(4, "P4"), Product(5, "P1"),
                    Product(6, "P1"), Product(7, "P2"), Product(8, "P3"), Product(9, "P4"), Product(10, "P1"),
                    Product(11, "P1"), Product(12, "P2"), Product(13, "P3"), Product(14, "P4"), Product(15, "P1")]"""
    # Creates an instance of type Robot with ID: 20
    mc_turner = Robot(20)

    ass_store = ["C1", "C2", "C3", "C4", "C5", "C6", "C5", "C5", "C4", "C1", "C1", "C2", "C3", "C4", "C5"]
    # ass_store = []
    # Creates an instance of type AssemblyLine with ID: 11 and Location: ??
    assembly_line = AssemblyLine(11, Location(position=[-1.584768513325685, -1.7801166514176172, 0.20401304075903764], orientation=[0.03795156580891418, 0.019484419185985095, 0.0832322780872719, 0.995616604896704]), ass_store)

    # Creates an instance of type QualityControl with ID: 33
    quality_control = QualityControl(33)

    # Adds to the robot storage for testing
    # mc_turner.add_to_storage("C1")
    # mc_turner.add_to_storage("C5")

    # Creates pickup spots for components
    component_spots = [PickupSpot(1, "C1", Location(position=[-0.616918207425653, -2.242842756135586, 0.19613864438951728], orientation=[-0.009431036312185292, 0.01231264555434063, 0.9922216495797277, -0.12351377420682844])), PickupSpot(2, "C2", Location(position=[-0.2526848031495088, -1.9472643882706038, 0.1913090855095235], orientation=[-0.006304768373071939, 0.03596877622757848, 0.9934187457888075, -0.10856192955103956])),
                       PickupSpot(3, "C3", Location(position=[0.3108487545847288, -1.5261153773619203, 0.18513323131825318], orientation=[-0.009759002245711364, 0.034387687922184586, 0.9944080981071959, -0.09937194379383027])), PickupSpot(4, "C4", Location(position=[-0.3481939223895538, -1.1952221455465557, 0.1943201055522587], orientation=[-0.006079794582326344, 0.03650010425249907, 0.9968041334066586, -0.07079758548703237])),
                       PickupSpot(5, "C5", Location(position=[0.40732846559848945, -2.089622572673389, 0.18586021116252788], orientation=[-0.015320796453234562, 0.03568879531847858, 0.992907332438075, -0.11236819957381244])), PickupSpot(6, "C6", Location(position=[-1.0030782814918058, -0.6323491507853269, 0.20239447521922718], orientation=[-0.0063303033380415945, 0.035794560507803705, 0.997557187986339, -0.05965176773985495]))]

    # Loads the interface
    main_window = load_interface(product_list, mc_turner, assembly_line, quality_control, component_spots)
    # Runs the main loop
    run_interface(main_window)
